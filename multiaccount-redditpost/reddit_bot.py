import praw
import config
import time
import os
import random
import openpyxl
import datetime
GonderilenPostIdleri = []
Satirsublar = []
Postlinks = []
Hata = "community"
import socket

Subreddit_adi="Teens_content_link"
        
def internet(host="8.8.8.8", port=53, timeout=3):
    """
    Host: 8.8.8.8 (google-public-dns-a.google.com)
    OpenPort: 53/tcp
    Service: domain (DNS/TCP)
    """
    try:
        socket.setdefaulttimeout(timeout)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except socket.error as ex:
        print(ex)
        return False

def randomTime():
    randomText1 = random.randint(120,150)
    return  randomText1


def bot_login():

    try:
        path = "users.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        for i in range(1, m_row+1):    
            username=sheet_obj.cell(row = i, column = 1).value
            password=sheet_obj.cell(row = i, column = 2).value
            client_id=sheet_obj.cell(row = i, column = 3).value
            client_secret=sheet_obj.cell(row = i, column = 4).value
            user_agent=sheet_obj.cell(row = i, column = 5).value
            with open("usernames.txt", "r") as file3:  #Paylasilansublar.txt elle oluştur.
                Usernames = file3.read()
                Usernames = Usernames.split("\n")
                Usernames = list(filter(None, Usernames))
                if username not in Usernames :
                    print ("Logging in..."+ "\n"+username+"\n")
                    r = praw.Reddit(username = username,
                                password = password,
                                client_id = client_id,
                                client_secret = client_secret,
                                user_agent = user_agent)
                    print ("Logged in!")
                    with open ("usernames.txt", "a") as f:  
                        f.write(username + "\n")
                    return r
    except Exception as e:    
        print(e)
        
    with open("usernames.txt", "r+") as file2:
        file2.truncate()
        print ("usernames.txt'in içi temizlendi.\n")


def run_bot(GonderilenPostIdleri):
    r=bot_login()
    try:
        subreddit1 = r.subreddit(Subreddit_adi)
    except Exception as e:
        r=bot_login()
        subreddit1 = r.subreddit(Subreddit_adi)
    for submission in subreddit1.new(limit=None):
        if submission.id not in GonderilenPostIdleri :
            print ("Yeni Post bulundu. Post id : " + submission.id)
     
            with open('subreddits.txt') as file:
                for line in file:
                    print("Subreddit sıraya alındı : " + line.rstrip())
                    with open("Paylasilansublar.txt") as file1:  #Paylasilansublar.txt elle oluştur.
                        Satirsublar = [line.strip() for line in file1]
                    if line.rstrip() not in Satirsublar :
                        submission1 = r.submission(id=submission.id)

                        try:

                              submission1.crosspost (subreddit=line.rstrip(), send_replies=True) # .reply(body=yorum)
       
                        except Exception as e:
                            print(str(e))
                            
                            if ("community" in str(e)  or  "flair"  in str(e)):
                                with open("subreddits.txt", "r") as f:
                                    lines = f.readlines()
                                with open("subreddits.txt", "w") as f:
                                    for line1 in lines:
                                        if (line1.strip("\n") != line.rstrip()):
                                            f.write(line1)
                                    pass 
                               
                                    
                            else:
                                 try:
                                      submission1.crosspost (subreddit=line.rstrip(), send_replies=True) #.reply(body=yorum)

                                 except:
                                    with open("subreddits.txt", "r") as f:
                                        lines = f.readlines()
                                    with open("subreddits.txt", "w") as f:
                                        for line1 in lines:
                                            if (line1.strip("\n") != line.rstrip()):
                                                f.write(line1)
                                        pass
          
                        with open ("Paylasilansublar.txt", "a") as f:
                            f.write(line.rstrip() + "\n")
                        print ("Post Göderimi "+line.rstrip()+" subredditine yapıldı :  " + submission.id+"\n")
                        print ("Crosspost için yeni subreddit bekleniyor (60sn) ", time.sleep(randomTime()))
            
            GonderilenPostIdleri.append(submission.id)
          
            with open ("GonderilenPostIdleri.txt", "a") as f:
                f.write(submission.id + "\n")
                print ("GonderilenPostIdleri.txt'nin içine crosspost yapılan id kaydedildi.")
            print ("Crosspost işlemiş Bitti. " + submission.id)
            with open("Paylasilansublar.txt", "r+") as file2:
                file2.truncate()
                print ("Paylasilansublar.txt'in içi temizlendi.\n")
                print ("Yeni içerik için crosspostu  2 saat beklet. ",time.sleep(7200),"\n") # crosspost bittikten 2 saat sonra tekrar başla.
            break
    print ("GONDERİLECEK POST KALMADI")
    time.sleep(10)
    
                
  
def get_saved_comments():

    if not os.path.isfile("GonderilenPostIdleri.txt"):
        GonderilenPostIdleri = []
    else:
        with open("GonderilenPostIdleri.txt", "r") as f:
            GonderilenPostIdleri = f.read()
            GonderilenPostIdleri = GonderilenPostIdleri.split("\n")
            GonderilenPostIdleri = list(filter(None, GonderilenPostIdleri))

    return GonderilenPostIdleri





GonderilenPostIdleri = get_saved_comments()

while True:

        run_bot(GonderilenPostIdleri)

    
        
